#!/usr/bin/env python3
"""Refresh Survival_Rescate in Datos_Graficas_Cox.xlsx from cumul_rescue_d100.dta."""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd


def main() -> int:
    outdir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("outputs")
    dta_path = outdir / "cumul_rescue_d100.dta"
    xlsx_path = outdir / "Datos_Graficas_Cox.xlsx"

    if not dta_path.exists():
        print(f"SKIP: {dta_path} not found")
        return 0

    df = pd.read_stata(dta_path)
    groups = ["FARC", "ELN", "Paramilitares", "Delincuencia"]
    surv_cols = ["surv1", "surv2", "surv3", "surv4"]

    rows = []
    for _, row in df.sort_values("_t").iterrows():
        rows.append(
            tuple(row[c] for c in surv_cols)
            + (float(row["_t"]),)
        )

    header = tuple(groups) + ("Dias",)
    sheet_rows = [header] + list(rows)

    try:
        import openpyxl
    except ImportError:
        print("WARN: openpyxl not installed; writing CSV fallback only")
        csv_path = outdir / "Survival_Rescate.csv"
        pd.DataFrame(sheet_rows[1:], columns=header).to_csv(csv_path, index=False)
        print(f"Wrote {csv_path}")
        return 0

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        if "Survival_Rescate" in wb.sheetnames:
            del wb["Survival_Rescate"]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet("Survival_Rescate")
    for r in sheet_rows:
        ws.append(list(r))

    tmax = df["_t"].max()
    idx = (df["_t"] - tmax).abs().idxmin()
    terminal = df.loc[idx]
    cumul = {groups[i]: round(1 - terminal[surv_cols[i]], 4) for i in range(4)}
    if "Cumul_Rescue" in wb.sheetnames:
        del wb["Cumul_Rescue"]
    ws2 = wb.create_sheet("Cumul_Rescue")
    ws2.append(["Grupo_Nombre", f"Pr_Rescue_day_{int(tmax)}"])
    for name, val in cumul.items():
        ws2.append([name, val])

    wb.save(xlsx_path)
    print(f"Updated {xlsx_path} (Survival_Rescate, Cumul_Rescue at day {int(tmax)})")
    for name, val in cumul.items():
        print(f"  {name}: {val}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
